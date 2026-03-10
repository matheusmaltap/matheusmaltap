"""
Planilha de Controle Financeiro
Gerador de planilha Excel para controle de receitas e despesas mensais.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint
import datetime

# ---------------------------------------------------------------------------
# Cores e estilos
# ---------------------------------------------------------------------------
COR_HEADER = "1F4E79"          # azul escuro
COR_RECEITA = "1E8449"         # verde
COR_DESPESA = "C0392B"         # vermelho
COR_SALDO_POS = "D5F5E3"       # verde claro
COR_SALDO_NEG = "FADBD8"       # vermelho claro
COR_TITULO_DASH = "2E86C1"     # azul médio
COR_ZEBRA = "EBF5FB"           # azul muito claro (linhas alternadas)
COR_BRANCO = "FFFFFF"

FONTE_TITULO = Font(name="Calibri", bold=True, size=14, color=COR_BRANCO)
FONTE_HEADER = Font(name="Calibri", bold=True, size=11, color=COR_BRANCO)
FONTE_DADO = Font(name="Calibri", size=11)
FONTE_TOTAL = Font(name="Calibri", bold=True, size=11)
FONTE_SALDO = Font(name="Calibri", bold=True, size=12)

BORDA_FINA = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

MESES = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]

CATEGORIAS_RECEITA = [
    "Salário",
    "Freelance / Renda extra",
    "Investimentos",
    "Outras receitas",
]

CATEGORIAS_DESPESA = [
    "Moradia (aluguel / financiamento)",
    "Alimentação",
    "Transporte",
    "Saúde",
    "Educação",
    "Lazer e entretenimento",
    "Vestuário",
    "Serviços (internet, telefone, streaming)",
    "Seguros",
    "Outras despesas",
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _apply_header_style(cell, bg_color: str, font: Font = FONTE_HEADER):
    cell.fill = _fill(bg_color)
    cell.font = font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDA_FINA


def _apply_data_style(cell, zebra: bool = False, number_format: str = None):
    cell.font = FONTE_DADO
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDA_FINA
    if zebra:
        cell.fill = _fill(COR_ZEBRA)
    if number_format:
        cell.number_format = number_format


def _col(n: int) -> str:
    """Retorna a letra da coluna dado o índice (1-based)."""
    return get_column_letter(n)


# ---------------------------------------------------------------------------
# Aba de lançamentos mensais
# ---------------------------------------------------------------------------

def _criar_aba_mes(wb: openpyxl.Workbook, mes: str, ano: int):
    ws = wb.create_sheet(title=mes)

    # --- Título principal ---
    ws.merge_cells("A1:G1")
    titulo = ws["A1"]
    titulo.value = f"Controle Financeiro — {mes}/{ano}"
    titulo.font = Font(name="Calibri", bold=True, size=16, color=COR_BRANCO)
    titulo.fill = _fill(COR_HEADER)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # --- Seção RECEITAS ---
    ws.merge_cells("A3:G3")
    sec = ws["A3"]
    sec.value = "📈  RECEITAS"
    sec.font = FONTE_TITULO
    sec.fill = _fill(COR_RECEITA)
    sec.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[3].height = 24

    headers_rec = ["Categoria", "Descrição", "Data", "Valor Previsto (R$)", "Valor Realizado (R$)", "Diferença (R$)", "Observação"]
    for col_idx, h in enumerate(headers_rec, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        _apply_header_style(cell, COR_RECEITA)
    ws.row_dimensions[4].height = 36

    row = 5
    for i, cat in enumerate(CATEGORIAS_RECEITA):
        zebra = (i % 2 == 0)
        ws.cell(row=row, column=1, value=cat)
        for col_idx in range(1, 8):
            c = ws.cell(row=row, column=col_idx)
            _apply_data_style(c, zebra=zebra)
        ws.cell(row=row, column=1).value = cat
        for col_idx in [4, 5]:
            ws.cell(row=row, column=col_idx).number_format = 'R$ #,##0.00'
        ws.cell(row=row, column=6).number_format = 'R$ #,##0.00'
        # Fórmula: Diferença = Realizado - Previsto
        ws.cell(row=row, column=6).value = f"=E{row}-D{row}"
        row += 1

    # Linha de total receitas
    tot_row_rec = row
    ws.merge_cells(f"A{tot_row_rec}:C{tot_row_rec}")
    t = ws.cell(row=tot_row_rec, column=1, value="TOTAL RECEITAS")
    t.font = FONTE_TOTAL
    t.fill = _fill("D5F5E3")
    t.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    t.border = BORDA_FINA
    ws.cell(row=tot_row_rec, column=2).border = BORDA_FINA
    ws.cell(row=tot_row_rec, column=3).border = BORDA_FINA

    for col_idx, col_letter_src in [(4, "D"), (5, "E"), (6, "F")]:
        c = ws.cell(row=tot_row_rec, column=col_idx)
        c.value = f"=SUM({col_letter_src}5:{col_letter_src}{tot_row_rec - 1})"
        c.font = FONTE_TOTAL
        c.fill = _fill("D5F5E3")
        c.number_format = 'R$ #,##0.00'
        c.border = BORDA_FINA

    row = tot_row_rec + 2

    # --- Seção DESPESAS ---
    ws.merge_cells(f"A{row}:G{row}")
    sec2 = ws.cell(row=row, column=1, value="📉  DESPESAS")
    sec2.font = FONTE_TITULO
    sec2.fill = _fill(COR_DESPESA)
    sec2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24

    headers_desp = ["Categoria", "Descrição", "Data", "Valor Previsto (R$)", "Valor Realizado (R$)", "Diferença (R$)", "Observação"]
    row += 1
    for col_idx, h in enumerate(headers_desp, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        _apply_header_style(cell, COR_DESPESA)
    ws.row_dimensions[row].height = 36

    row += 1
    desp_start = row
    for i, cat in enumerate(CATEGORIAS_DESPESA):
        zebra = (i % 2 == 0)
        for col_idx in range(1, 8):
            c = ws.cell(row=row, column=col_idx)
            _apply_data_style(c, zebra=zebra)
        ws.cell(row=row, column=1).value = cat
        for col_idx in [4, 5]:
            ws.cell(row=row, column=col_idx).number_format = 'R$ #,##0.00'
        ws.cell(row=row, column=6).number_format = 'R$ #,##0.00'
        ws.cell(row=row, column=6).value = f"=E{row}-D{row}"
        row += 1

    # Linha de total despesas
    tot_row_desp = row
    ws.merge_cells(f"A{tot_row_desp}:C{tot_row_desp}")
    t2 = ws.cell(row=tot_row_desp, column=1, value="TOTAL DESPESAS")
    t2.font = FONTE_TOTAL
    t2.fill = _fill("FADBD8")
    t2.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    t2.border = BORDA_FINA
    ws.cell(row=tot_row_desp, column=2).border = BORDA_FINA
    ws.cell(row=tot_row_desp, column=3).border = BORDA_FINA

    for col_idx, col_letter_src in [(4, "D"), (5, "E"), (6, "F")]:
        c = ws.cell(row=tot_row_desp, column=col_idx)
        c.value = f"=SUM({col_letter_src}{desp_start}:{col_letter_src}{tot_row_desp - 1})"
        c.font = FONTE_TOTAL
        c.fill = _fill("FADBD8")
        c.number_format = 'R$ #,##0.00'
        c.border = BORDA_FINA

    row = tot_row_desp + 2

    # --- Seção SALDO ---
    ws.merge_cells(f"A{row}:G{row}")
    sec3 = ws.cell(row=row, column=1, value="💰  SALDO DO MÊS")
    sec3.font = FONTE_TITULO
    sec3.fill = _fill(COR_HEADER)
    sec3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 28
    row += 1

    labels = [
        ("Saldo Previsto (Receitas - Despesas)", f"=D{tot_row_rec}-D{tot_row_desp}"),
        ("Saldo Realizado (Receitas - Despesas)", f"=E{tot_row_rec}-E{tot_row_desp}"),
        ("Variação (Realizado - Previsto)", f"=F{tot_row_rec}-F{tot_row_desp}"),
    ]
    for label, formula in labels:
        ws.merge_cells(f"A{row}:C{row}")
        lbl = ws.cell(row=row, column=1, value=label)
        lbl.font = FONTE_TOTAL
        lbl.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        lbl.border = BORDA_FINA
        ws.cell(row=row, column=2).border = BORDA_FINA
        ws.cell(row=row, column=3).border = BORDA_FINA
        val = ws.cell(row=row, column=4, value=formula)
        val.font = FONTE_SALDO
        val.number_format = 'R$ #,##0.00'
        val.border = BORDA_FINA
        for col_idx in range(5, 8):
            ws.cell(row=row, column=col_idx).border = BORDA_FINA
        row += 1

    # --- Largura das colunas ---
    col_widths = [30, 28, 14, 22, 22, 22, 28]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[_col(i)].width = w

    # --- Congelar painel abaixo do cabeçalho ---
    ws.freeze_panes = "A2"

    return ws


# ---------------------------------------------------------------------------
# Aba de resumo anual (Dashboard)
# ---------------------------------------------------------------------------

def _criar_aba_dashboard(wb: openpyxl.Workbook, ano: int):
    ws = wb.create_sheet(title="📊 Resumo Anual")
    wb.move_sheet(ws, offset=-len(wb.worksheets) + 1)  # move para a primeira posição

    # Título
    ws.merge_cells("A1:N1")
    t = ws["A1"]
    t.value = f"📊  RESUMO FINANCEIRO ANUAL — {ano}"
    t.font = Font(name="Calibri", bold=True, size=18, color=COR_BRANCO)
    t.fill = _fill(COR_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Legenda
    ws.merge_cells("A3:N3")
    leg = ws["A3"]
    leg.value = "Este resumo é calculado automaticamente a partir dos dados inseridos em cada aba mensal."
    leg.font = Font(name="Calibri", italic=True, size=11, color="555555")
    leg.alignment = Alignment(horizontal="center")

    # --- Tabela de resumo por mês ---
    headers = ["Mês", "Receita Prevista", "Receita Realizada", "Despesa Prevista", "Despesa Realizada", "Saldo Previsto", "Saldo Realizado"]
    row = 5
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        _apply_header_style(cell, COR_TITULO_DASH)
    ws.row_dimensions[row].height = 36

    row = 6
    for i, mes in enumerate(MESES):
        zebra = (i % 2 == 0)
        ws.cell(row=row, column=1, value=mes)
        # Índices: col D = Previsto Receita, col E = Realizado Receita (linha total_rec)
        # Fórmulas apontam para totais de cada aba mensal
        # Total receitas previstas = célula D{tot_row_rec} da aba
        # Usamos INDIRECT para referenciar dinamicamente
        ws.cell(row=row, column=2).value = f"=IFERROR(INDIRECT(\"'\" & A{row} & \"'!D\" & MATCH(\"TOTAL RECEITAS\",INDIRECT(\"'\" & A{row} & \"'!A:A\"),0)),0)"
        ws.cell(row=row, column=3).value = f"=IFERROR(INDIRECT(\"'\" & A{row} & \"'!E\" & MATCH(\"TOTAL RECEITAS\",INDIRECT(\"'\" & A{row} & \"'!A:A\"),0)),0)"
        ws.cell(row=row, column=4).value = f"=IFERROR(INDIRECT(\"'\" & A{row} & \"'!D\" & MATCH(\"TOTAL DESPESAS\",INDIRECT(\"'\" & A{row} & \"'!A:A\"),0)),0)"
        ws.cell(row=row, column=5).value = f"=IFERROR(INDIRECT(\"'\" & A{row} & \"'!E\" & MATCH(\"TOTAL DESPESAS\",INDIRECT(\"'\" & A{row} & \"'!A:A\"),0)),0)"
        ws.cell(row=row, column=6).value = f"=B{row}-D{row}"
        ws.cell(row=row, column=7).value = f"=C{row}-E{row}"

        for col_idx in range(1, 8):
            c = ws.cell(row=row, column=col_idx)
            c.border = BORDA_FINA
            c.font = FONTE_DADO
            c.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
            if col_idx > 1:
                c.number_format = 'R$ #,##0.00'
            if zebra:
                c.fill = _fill(COR_ZEBRA)
        row += 1

    # Linha de totais anuais
    tot_row = row
    ws.cell(row=tot_row, column=1, value="TOTAL ANUAL")
    ws.cell(row=tot_row, column=1).font = FONTE_TOTAL
    ws.cell(row=tot_row, column=1).fill = _fill("D6EAF8")
    ws.cell(row=tot_row, column=1).border = BORDA_FINA
    ws.cell(row=tot_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(2, 8):
        col_l = _col(col_idx)
        c = ws.cell(row=tot_row, column=col_idx, value=f"=SUM({col_l}6:{col_l}{tot_row - 1})")
        c.font = FONTE_TOTAL
        c.fill = _fill("D6EAF8")
        c.number_format = 'R$ #,##0.00'
        c.border = BORDA_FINA
        c.alignment = Alignment(horizontal="center", vertical="center")

    # --- Larguras de colunas ---
    col_widths = [16, 22, 22, 22, 22, 20, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[_col(i)].width = w

    # --- Gráfico de barras: receitas vs despesas realizadas por mês ---
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Receitas vs Despesas Realizadas por Mês"
    chart.y_axis.title = "Valor (R$)"
    chart.x_axis.title = "Mês"
    chart.style = 10
    chart.width = 22
    chart.height = 14

    data_rec = Reference(ws, min_col=3, min_row=5, max_row=5 + 12)
    data_desp = Reference(ws, min_col=5, min_row=5, max_row=5 + 12)
    cats = Reference(ws, min_col=1, min_row=6, max_row=5 + 12)

    chart.add_data(data_rec, titles_from_data=True)
    chart.add_data(data_desp, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = "1E8449"
    chart.series[1].graphicalProperties.solidFill = "C0392B"

    ws.add_chart(chart, "I5")

    # --- Gráfico de pizza: composição das despesas do total anual ---
    pie = PieChart()
    pie.title = "Composição das Despesas (Anual)"
    pie.style = 10
    pie.width = 14
    pie.height = 14

    # Para o gráfico de pizza das despesas, usa dados da aba Janeiro como exemplo
    # (o usuário pode customizar após gerar)
    ws.cell(row=tot_row + 3, column=1, value="Nota: Preencha os dados mensais para atualizar o resumo e os gráficos automaticamente.")
    ws.cell(row=tot_row + 3, column=1).font = Font(name="Calibri", italic=True, size=10, color="888888")
    ws.merge_cells(f"A{tot_row + 3}:G{tot_row + 3}")

    ws.freeze_panes = "A6"
    return ws


# ---------------------------------------------------------------------------
# Aba de categorias personalizáveis
# ---------------------------------------------------------------------------

def _criar_aba_categorias(wb: openpyxl.Workbook):
    ws = wb.create_sheet(title="⚙️ Categorias")

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = "⚙️  CATEGORIAS PERSONALIZÁVEIS"
    t.font = Font(name="Calibri", bold=True, size=14, color=COR_BRANCO)
    t.fill = _fill(COR_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A3:C3")
    leg = ws["A3"]
    leg.value = "Edite as categorias abaixo conforme sua necessidade. As abas mensais já possuem categorias padrão."
    leg.font = Font(name="Calibri", italic=True, size=10, color="555555")
    leg.alignment = Alignment(horizontal="left")

    # Receitas
    ws.cell(row=5, column=1, value="CATEGORIAS DE RECEITA").font = Font(bold=True, color=COR_BRANCO)
    ws.cell(row=5, column=1).fill = _fill(COR_RECEITA)
    ws.cell(row=5, column=1).border = BORDA_FINA
    ws.cell(row=5, column=1).alignment = Alignment(horizontal="center")
    ws.merge_cells("A5:B5")
    ws.cell(row=5, column=2).border = BORDA_FINA

    for i, cat in enumerate(CATEGORIAS_RECEITA, start=6):
        c = ws.cell(row=i, column=1, value=cat)
        c.border = BORDA_FINA
        c.font = FONTE_DADO
        if i % 2 == 0:
            c.fill = _fill(COR_ZEBRA)

    # Despesas
    start_desp = 6 + len(CATEGORIAS_RECEITA) + 1
    ws.cell(row=start_desp, column=1, value="CATEGORIAS DE DESPESA").font = Font(bold=True, color=COR_BRANCO)
    ws.cell(row=start_desp, column=1).fill = _fill(COR_DESPESA)
    ws.cell(row=start_desp, column=1).border = BORDA_FINA
    ws.cell(row=start_desp, column=1).alignment = Alignment(horizontal="center")
    ws.merge_cells(f"A{start_desp}:B{start_desp}")
    ws.cell(row=start_desp, column=2).border = BORDA_FINA

    for i, cat in enumerate(CATEGORIAS_DESPESA, start=start_desp + 1):
        c = ws.cell(row=i, column=1, value=cat)
        c.border = BORDA_FINA
        c.font = FONTE_DADO
        if i % 2 == 0:
            c.fill = _fill(COR_ZEBRA)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    return ws


# ---------------------------------------------------------------------------
# Aba de instruções
# ---------------------------------------------------------------------------

def _criar_aba_instrucoes(wb: openpyxl.Workbook):
    ws = wb.create_sheet(title="📋 Instruções")

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "📋  INSTRUÇÕES DE USO"
    t.font = Font(name="Calibri", bold=True, size=16, color=COR_BRANCO)
    t.fill = _fill(COR_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    instrucoes = [
        ("1. Navegação", "Use as abas na parte inferior para navegar entre os meses e o resumo anual."),
        ("2. Receitas", "Preencha os valores previstos e realizados de receita em cada aba mensal. A diferença é calculada automaticamente."),
        ("3. Despesas", "Preencha os valores previstos e realizados de despesa. As categorias estão pré-definidas conforme uso comum."),
        ("4. Saldo", "O saldo previsto e realizado é calculado automaticamente ao final de cada aba mensal."),
        ("5. Resumo Anual", "A aba '📊 Resumo Anual' consolida todos os meses automaticamente via fórmulas."),
        ("6. Categorias", "Na aba '⚙️ Categorias' você pode personalizar as categorias de receita e despesa."),
        ("7. Gráficos", "Os gráficos no Resumo Anual são atualizados conforme você preenche os dados mensais."),
        ("8. Formato de valores", "Insira os valores numéricos sem o símbolo R$. Ex: 3500.00 para R$ 3.500,00."),
        ("9. Data", "Utilize o formato DD/MM/AAAA para datas."),
        ("10. Backup", "Salve o arquivo regularmente. Recomenda-se manter uma cópia mensal."),
    ]

    row = 3
    for titulo, desc in instrucoes:
        ws.cell(row=row, column=1, value=titulo).font = Font(bold=True, size=12, name="Calibri")
        ws.cell(row=row, column=1).alignment = Alignment(vertical="center")
        ws.merge_cells(f"B{row}:D{row}")
        ws.cell(row=row, column=2, value=desc).font = Font(size=11, name="Calibri")
        ws.cell(row=row, column=2).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 28
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    return ws


# ---------------------------------------------------------------------------
# Função principal
# ---------------------------------------------------------------------------

def gerar_planilha(ano: int = None, nome_arquivo: str = None) -> str:
    """
    Gera a planilha de controle financeiro.

    Args:
        ano: Ano para a planilha. Padrão: ano atual.
        nome_arquivo: Caminho do arquivo de saída (.xlsx). Padrão: planilha_controle_financeiro_{ano}.xlsx

    Returns:
        Caminho do arquivo gerado.
    """
    if ano is None:
        ano = datetime.date.today().year
    if nome_arquivo is None:
        nome_arquivo = f"planilha_controle_financeiro_{ano}.xlsx"

    wb = openpyxl.Workbook()
    # Remove a aba padrão
    wb.remove(wb.active)

    # Dashboard primeiro (vai ser movida para a frente)
    _criar_aba_dashboard(wb, ano)

    # Abas mensais
    for mes in MESES:
        _criar_aba_mes(wb, mes, ano)

    # Abas de suporte
    _criar_aba_categorias(wb)
    _criar_aba_instrucoes(wb)

    # Propriedades do arquivo
    wb.properties.title = f"Controle Financeiro {ano}"
    wb.properties.creator = "Matheus Malta"
    wb.properties.description = "Planilha de controle financeiro pessoal com resumo anual."

    wb.save(nome_arquivo)
    print(f"✅  Planilha gerada com sucesso: {nome_arquivo}")
    return nome_arquivo


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys
    import os

    ano = int(sys.argv[1]) if len(sys.argv) > 1 else datetime.date.today().year
    saida = sys.argv[2] if len(sys.argv) > 2 else None
    caminho = gerar_planilha(ano=ano, nome_arquivo=saida)
    tamanho = os.path.getsize(caminho)
    print(f"📁  Arquivo: {caminho}  ({tamanho / 1024:.1f} KB)")
